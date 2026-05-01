from __future__ import annotations

from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..core.diff_engine import DiffResult

_FILL_A = PatternFill('solid', fgColor='FFD2D2')
_FILL_B = PatternFill('solid', fgColor='D2E6FF')
_FILL_MATCH = PatternFill('solid', fgColor='D2F5D2')
_FILL_HEADER = PatternFill('solid', fgColor='E0E0E0')
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal='center')


def _write_sheet(ws, headers: list[str], rows: list[dict], fill: PatternFill):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _BOLD
        cell.fill = _FILL_HEADER
        cell.alignment = _CENTER
    for row_dict in rows:
        row_data = [str(row_dict.get(h, '')) for h in headers]
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.fill = fill
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def write_diff(result: DiffResult, output_path: str, options: dict | None = None):
    if options is None:
        options = {'only_a': True, 'only_b': True, 'matched': True, 'summary': True}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if options.get('only_a') and result.only_in_a:
        ws = wb.create_sheet(f'Apenas em A')
        _write_sheet(ws, result.headers_a, result.only_in_a, _FILL_A)

    if options.get('only_b') and result.only_in_b:
        ws = wb.create_sheet(f'Apenas em B')
        _write_sheet(ws, result.headers_b, result.only_in_b, _FILL_B)

    if options.get('matched') and result.matched:
        ws = wb.create_sheet('Coincidências')
        headers_a = [f'A: {h}' for h in result.headers_a]
        headers_b = [f'B: {h}' for h in result.headers_b]
        headers = headers_a + headers_b + ['Similaridade %']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _BOLD
            cell.fill = _FILL_HEADER
            cell.alignment = _CENTER
        for pair in result.matched:
            row_a_vals = [str(pair.row_a.get(h, '')) for h in result.headers_a]
            row_b_vals = [str(pair.row_b.get(h, '')) for h in result.headers_b]
            ws.append(row_a_vals + row_b_vals + [f'{pair.score:.0f}%'])
            for cell in ws[ws.max_row]:
                cell.fill = _FILL_MATCH
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    if options.get('summary'):
        ws = wb.create_sheet('Resumo')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        rows = [
            ('Arquivo A', result.file_a_name),
            ('Arquivo B', result.file_b_name),
            ('Threshold de similaridade', f'{result.threshold}%'),
            ('Apenas em A', len(result.only_in_a)),
            ('Apenas em B', len(result.only_in_b)),
            ('Coincidências', len(result.matched)),
            ('Total de itens únicos', len(result.only_in_a) + len(result.only_in_b) + len(result.matched)),
            ('Gerado em', datetime.now().strftime('%d/%m/%Y %H:%M')),
        ]
        for label, value in rows:
            ws.append([label, value])
            ws[ws.max_row][0].font = _BOLD

    if not wb.sheetnames:
        ws = wb.create_sheet('Resultado vazio')
        ws['A1'] = 'Nenhum dado para exportar com as opções selecionadas.'

    wb.save(output_path)
