from __future__ import annotations

import re
import statistics
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from .loader import load_file

_NUMERIC_RE = re.compile(r'^\s*-?\d+([.,]\d+)?\s*$')


def _looks_numeric(values: list[str]) -> bool:
    non_empty = [v for v in values if v.strip()]
    if not non_empty:
        return False
    return sum(1 for v in non_empty if _NUMERIC_RE.match(v)) / len(non_empty) > 0.8


@dataclass
class ColumnSplitCandidate:
    column: str
    delimiter: str
    max_parts: int
    confidence: float
    proposed_names: list[str]
    enabled: bool = True


@dataclass
class SplitConfig:
    candidates: list[ColumnSplitCandidate]
    drop_original: bool = True


@dataclass
class SplitResult:
    headers: list[str]
    rows: list[dict]
    source_file: str
    applied_columns: list[str]


def detect_split_candidates(
    headers: list[str],
    rows: list[dict],
    min_confidence: float = 0.5,
) -> list[ColumnSplitCandidate]:
    candidates: list[ColumnSplitCandidate] = []

    for col in headers:
        values = [str(r.get(col, '')) for r in rows]
        non_empty = [v for v in values if v.strip()]
        if len(non_empty) < 2:
            continue
        if _looks_numeric(non_empty):
            continue

        count_semi = sum(1 for v in non_empty if ';' in v)
        count_comma = sum(1 for v in non_empty if ',' in v)
        total = len(non_empty)

        conf_semi = count_semi / total
        conf_comma = count_comma / total

        if conf_semi >= min_confidence:
            delimiter, confidence = ';', conf_semi
        elif conf_comma >= min_confidence and conf_comma > conf_semi + 0.2:
            delimiter, confidence = ',', conf_comma
        else:
            continue

        part_counts = [len(v.split(delimiter)) for v in non_empty if delimiter in v]
        if not part_counts:
            continue

        try:
            mode_parts = statistics.mode(part_counts)
        except statistics.StatisticsError:
            mode_parts = part_counts[0]

        if mode_parts < 2:
            continue

        try:
            std = statistics.stdev(part_counts) if len(part_counts) > 1 else 0.0
        except statistics.StatisticsError:
            std = 0.0
        if std > 2.0:
            continue

        proposed = [f'{col}_{i+1}' for i in range(mode_parts)]
        candidates.append(ColumnSplitCandidate(
            column=col,
            delimiter=delimiter,
            max_parts=mode_parts,
            confidence=confidence,
            proposed_names=proposed,
        ))

    return sorted(candidates, key=lambda c: c.confidence, reverse=True)


def _apply_split_to_rows(
    headers: list[str],
    rows: list[dict],
    config: SplitConfig,
) -> tuple[list[str], list[dict]]:
    enabled = {c.column: c for c in config.candidates if c.enabled}

    new_headers: list[str] = []
    for h in headers:
        if h in enabled:
            new_headers.extend(enabled[h].proposed_names)
        else:
            new_headers.append(h)

    new_rows: list[dict] = []
    for row in rows:
        new_row: dict = {}
        for h in headers:
            if h in enabled:
                cand = enabled[h]
                raw = str(row.get(h, ''))
                parts = raw.split(cand.delimiter)
                for i, name in enumerate(cand.proposed_names):
                    new_row[name] = parts[i].strip() if i < len(parts) else ''
            else:
                new_row[h] = row.get(h, '')
        new_rows.append(new_row)

    return new_headers, new_rows


def run_split(path: str | Path, config: SplitConfig) -> SplitResult:
    headers, rows = load_file(path)
    new_headers, new_rows = _apply_split_to_rows(headers, rows, config)
    applied = [c.column for c in config.candidates if c.enabled]
    return SplitResult(
        headers=new_headers,
        rows=new_rows,
        source_file=Path(path).name,
        applied_columns=applied,
    )


def write_split(result: SplitResult, output_path: str | Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Separado'

    header_fill = PatternFill('solid', fgColor='1565C0')
    header_font = Font(bold=True, color='FFFFFF')

    ws.append(result.headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in result.rows:
        ws.append([str(row.get(h, '')) for h in result.headers])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
