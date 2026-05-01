from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .loader import load_file


def _norm(text: str) -> str:
    t = str(text).lower().strip()
    nfd = unicodedata.normalize('NFD', t)
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')


@dataclass
class ColumnMapping:
    col_a: str | None
    col_b: str | None
    output_name: str


@dataclass
class MergeConfig:
    mappings: list[ColumnMapping]
    remove_duplicates: bool = False
    key_column: str = ''


@dataclass
class MergeResult:
    headers: list[str]
    rows: list[dict]
    rows_from_a: int
    rows_from_b: int
    source_a: str
    source_b: str


def detect_column_mappings(
    headers_a: list[str],
    headers_b: list[str],
) -> list[ColumnMapping]:
    norm_a = {_norm(h): h for h in headers_a}
    norm_b = {_norm(h): h for h in headers_b}

    mappings: list[ColumnMapping] = []
    used_b: set[str] = set()

    for norm, orig_a in norm_a.items():
        if norm in norm_b:
            orig_b = norm_b[norm]
            mappings.append(ColumnMapping(col_a=orig_a, col_b=orig_b, output_name=orig_a))
            used_b.add(norm)
        else:
            mappings.append(ColumnMapping(col_a=orig_a, col_b=None, output_name=orig_a))

    for norm, orig_b in norm_b.items():
        if norm not in used_b:
            mappings.append(ColumnMapping(col_a=None, col_b=orig_b, output_name=orig_b))

    return mappings


def run_merge(
    path_a: str | Path,
    path_b: str | Path,
    config: MergeConfig,
) -> MergeResult:
    _, rows_a = load_file(path_a)
    _, rows_b = load_file(path_b)

    output_names = [m.output_name for m in config.mappings]
    seen: set[str] = set()
    unique_names: list[str] = []
    for name in output_names:
        candidate = name
        i = 1
        while candidate in seen:
            candidate = f'{name}_{i}'
            i += 1
        seen.add(candidate)
        unique_names.append(candidate)

    for mapping, uname in zip(config.mappings, unique_names):
        mapping.output_name = uname

    def build_row(source: dict, side: str) -> dict:
        row: dict = {}
        for m in config.mappings:
            col = m.col_a if side == 'a' else m.col_b
            row[m.output_name] = str(source.get(col, '')) if col else ''
        return row

    merged: list[dict] = []
    for r in rows_a:
        merged.append(build_row(r, 'a'))
    for r in rows_b:
        merged.append(build_row(r, 'b'))

    if config.remove_duplicates and config.key_column:
        seen_keys: set[str] = set()
        deduped: list[dict] = []
        for row in merged:
            key = _norm(str(row.get(config.key_column, '')))
            if key not in seen_keys:
                seen_keys.add(key)
                deduped.append(row)
        merged = deduped

    return MergeResult(
        headers=unique_names,
        rows=merged,
        rows_from_a=len(rows_a),
        rows_from_b=len(rows_b),
        source_a=Path(path_a).name,
        source_b=Path(path_b).name,
    )


def write_merge(result: MergeResult, output_path: str | Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mesclado'

    header_fill = PatternFill('solid', fgColor='2E7D32')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    ws.append(result.headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in result.rows:
        ws.append([str(row.get(h, '')) for h in result.headers])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
